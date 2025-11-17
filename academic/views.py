from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Count, Avg, Q
from rest_framework import viewsets, status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import Docente, Curso, Estudiante, Actividad
from .serializers import (
    DocenteSerializer, CursoSerializer, 
    EstudianteSerializer, ActividadSerializer
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# ==========================================
# PÁGINA PRINCIPAL
# ==========================================
def index(request):
    """Renderiza el index.html"""
    return render(request, 'index.html')


# ==========================================
# API VIEWSETS (CRUD AUTOMÁTICO)
# ==========================================

class CursoViewSet(viewsets.ModelViewSet):
    """
    API para CRUD completo de Cursos
    GET    /api/cursos/          - Lista todos
    POST   /api/cursos/          - Crea nuevo
    GET    /api/cursos/1/        - Detalle
    PUT    /api/cursos/1/        - Actualiza
    DELETE /api/cursos/1/        - Elimina
    """
    queryset = Curso.objects.all()
    serializer_class = CursoSerializer
    
    def list(self, request):
        """Override para agregar estadísticas"""
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class EstudianteViewSet(viewsets.ModelViewSet):
    """API para CRUD completo de Estudiantes"""
    queryset = Estudiante.objects.all()
    serializer_class = EstudianteSerializer
    
    def retrieve(self, request, pk=None):
        """Obtener un estudiante específico"""
        try:
            estudiante = Estudiante.objects.get(pk=pk)
            serializer = self.get_serializer(estudiante)
            return Response(serializer.data)
        except Estudiante.DoesNotExist:
            return Response(
                {'error': 'Estudiante no encontrado'}, 
                status=status.HTTP_404_NOT_FOUND
            )


class ActividadViewSet(viewsets.ModelViewSet):
    """API para CRUD completo de Actividades"""
    queryset = Actividad.objects.all()
    serializer_class = ActividadSerializer
    
    def get_queryset(self):
        """Permite filtrar por curso con ?id_curso=1"""
        queryset = Actividad.objects.all()
        id_curso = self.request.query_params.get('id_curso', None)
        
        if id_curso:
            queryset = queryset.filter(id_curso=id_curso)
        
        return queryset.order_by('-fecha_entrega')
    
    def retrieve(self, request, pk=None):
        """Obtener una actividad específica"""
        try:
            actividad = Actividad.objects.get(pk=pk)
            serializer = self.get_serializer(actividad)
            return Response(serializer.data)
        except Actividad.DoesNotExist:
            return Response(
                {'error': 'Actividad no encontrada'}, 
                status=status.HTTP_404_NOT_FOUND
            )


# ==========================================
# API REPORTES
# ==========================================

@api_view(['GET'])
def reportes(request):
    """
    Endpoint de reportes con múltiples acciones
    GET /api/reportes/?action=general
    GET /api/reportes/?action=cursos_estadisticas
    GET /api/reportes/?action=estudiantes_por_curso
    GET /api/reportes/?action=rendimiento
    GET /api/reportes/?action=actividades_pendientes
    GET /api/reportes/?action=top_estudiantes
    GET /api/reportes/?action=promedios_mensuales
    """
    action = request.GET.get('action', 'general')
    
    try:
        if action == 'general':
            # Estadísticas generales del sistema
            cursos_activos = Curso.objects.filter(estado='Activo').count()
            total_estudiantes = Estudiante.objects.count()
            total_actividades = Actividad.objects.count()
            
            # Promedio general
            estudiantes_con_nota = Estudiante.objects.filter(nota_final__isnull=False)
            if estudiantes_con_nota.exists():
                promedio_general = estudiantes_con_nota.aggregate(Avg('nota_final'))['nota_final__avg']
                promedio_general = round(float(promedio_general), 2) if promedio_general else 0.0
            else:
                promedio_general = 0.0
            
            data = {
                'cursos_activos': cursos_activos,
                'total_estudiantes': total_estudiantes,
                'total_actividades': total_actividades,
                'promedio_general': promedio_general
            }
            return Response(data)
        
        elif action == 'cursos_estadisticas':
            # Estadísticas detalladas por curso
            cursos = Curso.objects.filter(estado='Activo')
            resultados = []
            
            for curso in cursos:
                num_estudiantes = Estudiante.objects.filter(id_curso=curso).count()
                num_actividades = Actividad.objects.filter(id_curso=curso).count()
                
                # Calcular promedio
                estudiantes_con_nota = Estudiante.objects.filter(
                    id_curso=curso, 
                    nota_final__isnull=False
                )
                
                if estudiantes_con_nota.exists():
                    promedio = estudiantes_con_nota.aggregate(Avg('nota_final'))['nota_final__avg']
                    promedio = round(float(promedio), 2) if promedio else None
                else:
                    promedio = None
                
                resultados.append({
                    'id_curso': curso.id_curso,
                    'nombre': curso.nombre,
                    'codigo': curso.codigo,
                    'num_estudiantes': num_estudiantes,
                    'num_actividades': num_actividades,
                    'promedio': promedio
                })
            
            # Ordenar por promedio descendente
            resultados.sort(key=lambda x: x['promedio'] if x['promedio'] else 0, reverse=True)
            return Response(resultados)
        
        elif action == 'estudiantes_por_curso':
            # Distribución de estudiantes por curso (Top 10)
            cursos = Curso.objects.filter(estado='Activo').annotate(
                cantidad=Count('estudiante')
            ).values('nombre', 'cantidad').order_by('-cantidad')[:10]
            
            resultados = [
                {'curso': c['nombre'], 'cantidad': c['cantidad']} 
                for c in cursos
            ]
            return Response(resultados)
        
        elif action == 'rendimiento':
            # Distribución de rendimiento (Aprobados/Reprobados/Sin Calificar)
            estudiantes = Estudiante.objects.all()
            
            aprobados = estudiantes.filter(nota_final__gte=3.0).count()
            reprobados = estudiantes.filter(nota_final__lt=3.0, nota_final__isnull=False).count()
            sin_calificar = estudiantes.filter(nota_final__isnull=True).count()
            
            return Response([
                {'estado': 'Aprobados', 'cantidad': aprobados},
                {'estado': 'Reprobados', 'cantidad': reprobados},
                {'estado': 'Sin Calificar', 'cantidad': sin_calificar}
            ])
        
        elif action == 'actividades_pendientes':
            # Actividades pendientes por curso
            from django.utils import timezone
            
            actividades_pendientes = Actividad.objects.filter(
                Q(estado='Pendiente') | Q(fecha_entrega__gt=timezone.now().date())
            ).values('id_curso__nombre').annotate(
                pendientes=Count('id_actividad')
            ).order_by('-pendientes')
            
            resultados = [
                {
                    'curso': a['id_curso__nombre'] or 'Sin curso', 
                    'pendientes': a['pendientes']
                } 
                for a in actividades_pendientes
            ]
            return Response(resultados)
        
        elif action == 'top_estudiantes':
            # Top 10 mejores estudiantes
            top_estudiantes = Estudiante.objects.filter(
                nota_final__isnull=False
            ).select_related('id_curso').order_by('-nota_final')[:10]
            
            resultados = [
                {
                    'estudiante': e.nombre,
                    'curso': e.id_curso.nombre if e.id_curso else 'Sin curso',
                    'promedio': float(e.nota_final)
                }
                for e in top_estudiantes
            ]
            return Response(resultados)
        
        elif action == 'promedios_mensuales':
            # Promedios mensuales (simulado)
            from django.db.models.functions import TruncMonth
            
            # Agrupar actividades por mes
            actividades_por_mes = Actividad.objects.filter(
                fecha_entrega__isnull=False
            ).values('fecha_entrega__year', 'fecha_entrega__month').distinct()
            
            resultados = []
            for item in actividades_por_mes:
                year = item['fecha_entrega__year']
                month = item['fecha_entrega__month']
                mes_str = f"{year}-{month:02d}"
                
                # Calcular promedio de estudiantes en ese mes
                estudiantes_mes = Estudiante.objects.filter(
                    nota_final__isnull=False
                )
                
                if estudiantes_mes.exists():
                    promedio = estudiantes_mes.aggregate(Avg('nota_final'))['nota_final__avg']
                    promedio = round(float(promedio), 2) if promedio else 0
                else:
                    promedio = 0
                
                resultados.append({
                    'mes': mes_str,
                    'promedio': promedio
                })
            
            resultados.sort(key=lambda x: x['mes'])
            return Response(resultados)
        
        else:
            return Response(
                {'success': False, 'error': 'Acción no válida'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
    
    except Exception as e:
        return Response(
            {'success': False, 'error': str(e)}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ==========================================
# EXPORTACIÓN A EXCEL
# ==========================================

@api_view(['GET'])
def exportar_excel(request):
    """
    Exporta datos a Excel
    GET /api/exportar/?tipo=estudiantes
    GET /api/exportar/?tipo=cursos
    GET /api/exportar/?tipo=actividades
    GET /api/exportar/?tipo=reporte_completo
    GET /api/exportar/?tipo=estudiantes&id_curso=1
    """
    tipo = request.GET.get('tipo', 'estudiantes')
    id_curso = request.GET.get('id_curso', None)
    
    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    
    # Estilos para encabezados
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    try:
        if tipo == 'estudiantes':
            ws.title = "Estudiantes"
            headers = ['ID', 'Nombre Completo', 'Curso', 'Código Curso', 'Nota Final', 'Estado']
            ws.append(headers)
            
            # Aplicar estilos
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Obtener datos
            estudiantes = Estudiante.objects.select_related('id_curso')
            if id_curso:
                estudiantes = estudiantes.filter(id_curso=id_curso)
            
            for e in estudiantes:
                ws.append([
                    e.id_estudiante,
                    e.nombre or '',
                    e.id_curso.nombre if e.id_curso else 'Sin curso',
                    e.id_curso.codigo if e.id_curso else '-',
                    float(e.nota_final) if e.nota_final else '-',
                    e.estado
                ])
        
        elif tipo == 'cursos':
            ws.title = "Cursos"
            headers = ['ID', 'Nombre', 'Código', 'Descripción', 'Estado', 
                      'Docente', 'Total Estudiantes', 'Total Actividades', 'Promedio']
            ws.append(headers)
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            cursos = Curso.objects.select_related('id_docente')
            
            for c in cursos:
                num_estudiantes = Estudiante.objects.filter(id_curso=c).count()
                num_actividades = Actividad.objects.filter(id_curso=c).count()
                
                estudiantes_con_nota = Estudiante.objects.filter(id_curso=c, nota_final__isnull=False)
                if estudiantes_con_nota.exists():
                    promedio = estudiantes_con_nota.aggregate(Avg('nota_final'))['nota_final__avg']
                    promedio = round(float(promedio), 2) if promedio else '-'
                else:
                    promedio = '-'
                
                ws.append([
                    c.id_curso,
                    c.nombre or '',
                    c.codigo or '',
                    c.descripcion or '',
                    c.estado or '',
                    c.id_docente.nombre if c.id_docente else 'Sin docente',
                    num_estudiantes,
                    num_actividades,
                    promedio
                ])
        
        elif tipo == 'actividades':
            ws.title = "Actividades"
            headers = ['ID', 'Nombre', 'Tipo', 'Curso', 'Código Curso', 
                      'Fecha Entrega', 'Porcentaje (%)', 'Estado']
            ws.append(headers)
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            actividades = Actividad.objects.select_related('id_curso')
            if id_curso:
                actividades = actividades.filter(id_curso=id_curso)
            
            for a in actividades:
                ws.append([
                    a.id_actividad,
                    a.nombre or '',
                    a.tipo or '',
                    a.id_curso.nombre if a.id_curso else 'Sin curso',
                    a.id_curso.codigo if a.id_curso else '-',
                    a.fecha_entrega.strftime('%d/%m/%Y') if a.fecha_entrega else '-',
                    a.porcentaje or 0,
                    a.estado or ''
                ])
        
        elif tipo == 'reporte_completo':
            ws.title = "Reporte Completo"
            headers = ['Curso', 'Código', 'Estudiante', 'Nota Final', 
                      'Estado', 'Actividades del Curso']
            ws.append(headers)
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            estudiantes = Estudiante.objects.filter(
                id_curso__estado='Activo'
            ).select_related('id_curso')
            
            for e in estudiantes:
                num_actividades = Actividad.objects.filter(id_curso=e.id_curso).count() if e.id_curso else 0
                
                ws.append([
                    e.id_curso.nombre if e.id_curso else 'Sin curso',
                    e.id_curso.codigo if e.id_curso else '-',
                    e.nombre or '',
                    float(e.nota_final) if e.nota_final else '-',
                    e.estado,
                    num_actividades
                ])
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Agregar pie de página con información
        ws.append([])
        ws.append(['Reporte generado:', datetime.now().strftime('%d/%m/%Y %H:%M:%S')])
        ws.append(['Sistema:', 'Plataforma de Gestión Académica - Django'])
        ws.append(['Total de registros:', len(list(ws.rows)) - 4])
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_{tipo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
    
    except Exception as e:
        return Response(
            {'success': False, 'error': str(e)}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )